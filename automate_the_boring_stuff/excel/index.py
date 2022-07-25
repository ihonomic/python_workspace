import openpyxl


wb = openpyxl.load_workbook('examples.xlsx')

sheet_names = wb.sheetnames
print(sheet_names, '<<SHEET NAMES<<<')

sheet1 = wb['Sheet1']
print(sheet1.title, '<<<TITLE OF SHEET<<')

print(sheet1.max_row, sheet1.max_column, '<<<Max Column and Row of a sheet<<')
#   Accessing cell
print(sheet1['A1'].value, '<<<Value OF cell<<')
print(sheet1['A1'].row, '<<<Row OF cell<<')
print(sheet1['B1'].column, '<<<Column OF cell<<')
print(sheet1['A1'].coordinate, '<<<Coordinate OF cell<<')

print(sheet1.cell(row=1, column=2).value, '<<<Value OF cell<<')

for xSection in sheet1['A1':'C4']:
    for cell in xSection:
        print(cell.coordinate, cell.value)
    print('---End of Row---')


wb.create_sheet(index=0, title='First Sheet')
wb.remove(wb['Sheet2'])
wb.save('examples.xlsx')
