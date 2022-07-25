import openpyxl

number = int(input("Enter a Number between 1-9: "))
wb = openpyxl.Workbook()
sheet = wb.create_sheet(title='Multiplication table', index=1)

for figure in range(1, number+1):
    sheet.cell(row=figure+1, column=1).value = figure
    sheet.cell(row=1, column=figure+1).value = figure


for val in range(2, number+2):
    rowValue = sheet.cell(row=val, column=1).value
    # print(rowValue)

    for col in range(2, number+2):
        colValue = sheet.cell(row=1, column=col).value
        # print(colValue, '---COL')
        prod = rowValue * colValue

        sheet.cell(row=val, column=col).value = prod


wb.save('./excel/productTable/multiply.xlsx')
