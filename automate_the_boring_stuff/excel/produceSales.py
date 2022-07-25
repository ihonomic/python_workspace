import openpyxl

productPrice = {
    'Garlic': 100,
    'Potatoes': 2,
    'Watermelon': 900
}

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

for rowNumber in range(2, sheet.max_row):
    product = sheet.cell(row=rowNumber, column=1).value
    if product in productPrice.keys():
        sheet.cell(row=rowNumber, column=2).value = productPrice[product]

print('---DONE---')
wb.save('UpdatedProductSales.xlsx')
