import bs4
import requests
import openpyxl
from pprint import pprint

# Download the page
url = 'https://play.google.com/store/apps/collection/cluster?clp=ogoeCBEqAggIMhYKEHJ1LmlpZWMucHlkcm9pZDMQARgD:S:ANO1ljKU-j8&gsr=CiGiCh4IESoCCAgyFgoQcnUuaWllYy5weWRyb2lkMxABGAM%3D:S:ANO1ljI2N3s&hl=en&gl=US'
page = requests.get(url)
page.raise_for_status()

#   Beautiful soup
soup = bs4.BeautifulSoup(page.text, features='html.parser')

title = soup.select('.nnK0zc')
links = soup.select('.Q9MA7b a')

config = dict()

#   Write to excel sheel
wb = openpyxl.Workbook()
wb.create_sheet(title='Google App', index=0)
sheet = wb['Google App']
sheet['A1'].value = 'NAME OF APP'
sheet['B1'].value = 'LINKS'
wb.save('./scrapInfo/app_scrap.xlsx')

wb = openpyxl.load_workbook('./scrapInfo/app_scrap.xlsx')
sheet = wb['Google App']

for index in range(len(title)):
    name = title[index].getText()
    url = 'https://play.google.com' + links[index].get('href')

    config[name] = url

    sheet.cell(row=index+1, column=1).value = name
    sheet.cell(row=index+1, column=2).value = url

# pprint(config)

wb.save('./scrapInfo/app_scrap.xlsx')
print("----DONE----")
